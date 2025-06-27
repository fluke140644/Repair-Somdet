
from django.shortcuts import get_object_or_404, render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_POST
from .models import RepairRequest
from django.http import  HttpResponseForbidden, JsonResponse ,HttpResponse
from .forms import RepairRequestForm
import openpyxl
import json
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.contrib.auth.models import User 
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.db.models import Count
from openpyxl.utils import get_column_letter
from threading import Thread
from django.shortcuts import render, redirect
# pdf 
from weasyprint import HTML, CSS
import os
from django.template.loader import render_to_string
from django.conf import settings

# หน้าหลัก 
@login_required
def index(request):
    return render(request,"index.html")

def copy(request):
    return render(request,"copy.html")

def copy19(request):
    return render(request,"copy19.html")

@login_required
def login_view(request):
    return render(request, 'login.html')

@login_required
def base(request):
    return render(request,"base.html")

@login_required
def about(request):
    return render(request,"about.html")

@login_required
def pl_admin(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
    reports = RepairRequest.objects.all()
    for report in reports:
        if report.completed_at and report.in_progress_at:
            report.duration = report.completed_at - report.in_progress_at
        else:
            report.duration = None
    return render(request, 'pl_admin.html', {'reports': reports})



@login_required
def pl_adminlist(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
    return render(request,"pl_adminlist.html")



@login_required
def login_redirect(request):
    if request.user.is_staff:
        return redirect('dashadmin')
    else:
        return redirect('/')
    

class MyLoginView(LoginView):
    def get_success_url(self):
        if self.request.user.is_staff:
            return reverse('pl_admin')
        return reverse('index')


def send_repair_email_async(repair, user):
    def task():
        recipient_list = [admin.email for admin in User.objects.filter(is_staff=True) if admin.email]
        if not recipient_list:
            return

        context = {
            'user': user,
            'category': repair.category,
            'sub_item': repair.sub_item,
            'detail': repair.detail,
            'department': repair.department,
            'status_user': repair.status_user,
            'request_id': repair.request_id,
            'created_at': repair.created_at,
            'report': repair,
        }

        subject = '📥 มีรายการแจ้งซ่อมใหม่'
        text_content = f"""
รายการแจ้งซ่อมใหม่:

ผู้แจ้ง: {user.get_full_name()} ({user.username})
หมวดหมู่: {repair.category}
รายละเอียด: {repair.detail}
แผนก: {repair.department}
สถานะผู้แจ้ง: {repair.status_user}
"""

        html_content = render_to_string('repair_notification.html', context)

        msg = EmailMultiAlternatives(subject, text_content, None, recipient_list)
        msg.attach_alternative(html_content, "text/html")
        msg.send()

    Thread(target=task).start() 


@login_required
def create_repair_request(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        sub_item = request.POST.get('sub_item')
        detail = request.POST.get('description')
        department = request.POST.get('department')
        device_code = request.POST.get('device_code')
        unknown_device = request.POST.get('unknown_device') == 'on'
        status_user = request.POST.get('status_user')
        asset_number = request.POST.get('asset_number')

        if not (category and detail and department and status_user):
            messages.error(request, "กรุณากรอกข้อมูลให้ครบถ้วน")
            return render(request, 'index.html')

        repair = RepairRequest.objects.create(
            category=category,
            sub_item=sub_item or '',
            detail=detail,
            department=department,
            device_code=device_code or '',
            unknown_device=unknown_device,
            status_user=status_user,
            status='pending',
            created_by=request.user,
            reported_by=request.user,
            asset_number=asset_number or '',
        )

        send_repair_email_async(repair, request.user)

        messages.success(request, "ส่งคำขอแจ้งซ่อมเรียบร้อยแล้ว")
        return redirect('index')

    return render(request, 'index.html',{'report': repair,})

@login_required
def repair_list(request):
    reports = RepairRequest.objects.all()
    return render(request, 'repair_list.html', {'reports': reports})

@login_required
def repair_detail(request, pk):
    report = get_object_or_404(RepairRequest, pk=pk)
    return render(request, 'repair_detail.html', {'report': report})

@login_required
def repair_table(request):
    reports = RepairRequest.objects.all()
    return render(request, 'repair_table.html', {
        'reports': reports,
        'show_action_buttons': False, 
    })

@login_required
def pl_admin_view(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
    reports = RepairRequest.objects.all()
    return render(request, 'pl_admin.html', {'reports': reports})

#ปุ่มแก้ไข

@login_required
def repair_edit(request, id):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")

    repair = get_object_or_404(RepairRequest, id=id)
    original_repairclaim_at = repair.repairclaim_at

    if request.method == 'POST':
        form = RepairRequestForm(request.POST, instance=repair)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.repairclaim_at = original_repairclaim_at
            updated.save()
            return redirect('pl_admin')
    else:
        form = RepairRequestForm(instance=repair)

    return render(request, 'repair_edit.html', {'form': form, 'repair': repair})



@require_POST
@login_required
def repair_delete(request, id):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
    repair = get_object_or_404(RepairRequest, id=id)
    if request.user == repair.created_by or request.user.is_staff:
        repair.delete()
        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'forbidden'}, status=403)
    


@login_required
def update_status(request, id):
    if request.method == 'POST' and request.user.is_staff:
        try:
            repair = RepairRequest.objects.get(id=id)
            data = json.loads(request.body)
            new_status = data.get('status')

            if new_status not in dict(RepairRequest.STATUS_CHOICES):
                return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)

            repair.status = new_status
            repair.assigned_to = request.user

            if new_status == 'in_progress' and not repair.in_progress_at:
                repair.in_progress_at = now()
            if new_status == 'completed' and not repair.completed_at:
                repair.completed_at = now()
            if new_status == 'repairclaim' and not repair.repairclaim_at:
                repair.repairclaim_at = now()
                print("DEBUG: repairclaim_at set to", repair.repairclaim_at)

                

            repair.save()
            return JsonResponse({'success': True})
        except RepairRequest.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'ไม่พบรายการ'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON ไม่ถูกต้อง'}, status=400)
        

    return JsonResponse({'success': False, 'error': 'ไม่อนุญาต'}, status=403)


@login_required
def repair_list_view(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")
    reports = RepairRequest.objects.all()
    
    show_action_buttons = False
    if request.user.is_staff:
        show_action_buttons = True
    
    context = {
        'reports': reports,
        'show_action_buttons': show_action_buttons,
    }
    
    return render(request, 'repair_table_partial.html', context)

@login_required
def dashadmin(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงหน้านี้")

    pending_count = RepairRequest.objects.filter(status='pending').count()
    in_progress_count = RepairRequest.objects.filter(status='in_progress').count()
    completed_count = RepairRequest.objects.filter(status='completed').count()
    repairclaim_count = RepairRequest.objects.filter(status='repairclaim').count()
    cancelled_count = RepairRequest.objects.filter(status='cancelled').count()

    
    category_counts = RepairRequest.objects.values('category').annotate(total=Count('category')).order_by('-total')
    category_labels = [item['category'] for item in category_counts]
    category_values = [item['total'] for item in category_counts]

    all_reports = RepairRequest.objects.order_by('-created_at')
    paginator = Paginator(all_reports, 5)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'pending_count': pending_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
        'repairclaim': repairclaim_count,
        'cancelled_count': cancelled_count,
        'latest_reports': page_obj,
        'page_obj': page_obj,
        'category_labels': category_labels,   
        'category_values': category_values,   
    }

    return render(request, 'dashadmin.html', context)


@login_required
def my_repairs(request):
    user_reports = RepairRequest.objects.filter(reported_by=request.user).order_by('-created_at')
    return render(request, 'my_repairs.html', {'user_reports': user_reports})

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, 'เปลี่ยนรหัสผ่านเรียบร้อยแล้ว')
            return redirect('/')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'change_password.html', {'form': form})

def export_repair_requests_excel(request):
    repairs = RepairRequest.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Repair Requests"

    headers = [
        'ID', 'รหัสแจ้งซ่อม','ประเภท','หัวข้อย่อย','รหัสอุปกรณ์','รายละเอียด',
        'แผนก','สถานะ','ความเร่งด่วน','วันที่แจ้ง','ผู้แจ้ง (คนแรก)','ผู้แจ้งซ่อม',
        'วันที่เริ่มดำเนินการ','วันที่เสร็จสิ้น','วันที่ส่งซ่อม/เคลม','เจ้าหน้าที่รับแจ้ง',
    ]
    ws.append(headers)

    for repair in repairs:
        ws.append([
            repair.id,
            repair.request_id,
            repair.category,
            repair.sub_item,
            repair.device_code,
            repair.detail,
            repair.department,
            repair.status,
            repair.status_user,
            repair.created_at.strftime('%Y-%m-%d %H:%M') if repair.created_at else '',
            repair.created_by.username if repair.created_by else '',
            repair.reported_by.username if repair.reported_by else '',
            repair.in_progress_at.strftime('%Y-%m-%d %H:%M') if repair.in_progress_at else '',
            repair.completed_at.strftime('%Y-%m-%d %H:%M') if repair.completed_at else '',
            repair.repairclaim_at.strftime('%Y-%m-%d %H:%M') if repair.repairclaim_at else '',
            repair.assigned_to.username if repair.assigned_to else '',
        ])

    # ปรับความกว้างคอลัมน์
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # ส่งไฟล์
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=repair_requests.xlsx'
    wb.save(response)
    return response

# =============================== PDF ================================
from django.utils.dateparse import parse_date
from datetime import datetime

def monthly_report_pdf(request):
    
    month_str = request.GET.get('month')
    
    if not month_str:
        return HttpResponse("กรุณาระบุเดือน ?month=YYYY-MM", status=400)

    try:
        year, month = map(int, month_str.split('-'))
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
    except Exception:
        return HttpResponse("รูปแบบเดือนไม่ถูกต้อง", status=400)

    from .models import RepairRequest
    data = RepairRequest.objects.filter(created_at__gte=start_date, created_at__lt=end_date)

    thai_months = [
        '', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
        'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
    ]
    month_thai = f"{thai_months[month]} {year + 543}"

    context = {
        'month': month_thai,
        'data': data,  
    }

    html_string = render_to_string('monthly_report.html', context)

    font_path = os.path.join(settings.BASE_DIR, 'Report/static/Report/fonts/NotoSansThai-Regular.ttf')

    css_string = f'''
    @font-face {{
        font-family: 'NotoSansThai';
        src: url('file://{font_path}');
    }}
    body {{
        font-family: 'NotoSansThai', sans-serif;
        font-size: 10pt;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        margin-top: 12px;
    }}
    th, td {{
        border: 1px solid black;
        padding: 6px;
        font-size: 10pt;
    }}
    th {{
        background-color: #f2f2f2;
    }}
    h2 {{
        text-align: center;
    }}
    '''

    html = HTML(string=html_string)
    css = CSS(string=css_string)
    pdf_file = html.write_pdf(stylesheets=[css])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="แจ้งซ่อมประจำเดือน {{ month }}.pdf"'
    return response
